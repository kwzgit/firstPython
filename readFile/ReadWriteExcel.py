#!/usr/bin/python3
from openpyxl import load_workbook,Workbook
def connect_list_to_str(origin_list, cursor):
	#将列表以cursor为分隔符链接为字符串
	result = ''
	for obj in origin_list:
		result += cursor
		result += str(obj)
	if result != '':
		result = result[len(cursor):]
	return result
def connect_dict_to_str(origin_dict, key_cursor=':', dict_cursor='\n', list_cursor=','):
	result = ''
	for key,value in origin_dict.items():
		result += str(key) + str(key_cursor)
		if type(value) == list:
			result += connect_list_to_str(value, list_cursor)
		else:
			result += str(value)
		result += str(dict_cursor)
	result = result[:-len(dict_cursor)]
	return result
# 读Excel
def read_from_xl(filename, sheetname='Sheet'):
	# 从excel中读取信息，[{'dict':str},{},{}]
	dicts_list = []
	dicts_list_xl = load_workbook(filename + '.xlsx', data_only=True)
	Sheet = dicts_list_xl[sheetname]
	keys = []
	max_column = Sheet.max_column
	max_row = Sheet.max_row
	# print(str(max_column)+'列'+' '+str(max_row)+'行')
	for column in range(1, max_column+1):
		key = Sheet.cell(1,column).value
		keys.append(key)
	for row in range(2, max_row+1):
		dicts = {}
		for column in range(1, max_column+1):
			value = Sheet.cell(row,column).value
			if value is None:
				value = ''
			dicts[keys[column-1]] = value
		dicts_list.append(dicts)
	return dicts_list
def save_in_xl(dicts_list, filename, cut_cursor='\n', save='new', sheetname='Sheet'):
	# 存储到excel中，[{'dict':[list]},{},{}]
	if dicts_list == []: # 空文档
		print('\n--------------------------------------')
		print('save_in_xl: the list is empty !!\n')
	else:
		if save == 'new': # 存在新文件里
			xl = Workbook()
			sheet = xl.create_sheet(sheetname, 0)
		elif save == 'old': # 存在旧文件里
			xl = load_workbook(filename + '.xlsx')
			sheet = xl.create_sheet(sheetname, 0)
		# 存表头,默认第一行数据的key为所有数据的key
		sheet.append(list(dicts_list[0].keys()))
		for dicts in dicts_list:
			row = []
			for key,value in dicts.items(): # 转换单元格内的数据为字符串
				if type(value) == list:
					row.append(connect_list_to_str(value, cut_cursor))
				elif type(value) == dict:
					row.append(connect_dict_to_str(value, key_cursor='::', dict_cursor='\n', list_cursor=' | '))
				else:
					row.append(str(value))
			sheet.append(row) # 存内容
		xl.save(filename + '.xlsx')


if __name__ == "__main__":
	print()

	# save_in_xl(newDictList,'./excelFile/辅检图谱疾病',cut_cursor='\n',save='old',sheetname='胆管结石')



